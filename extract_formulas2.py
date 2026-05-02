"""Print the column header row and first Cu data row formulas for cross-check."""
import openpyxl, warnings
warnings.filterwarnings("ignore")

wb_f = openpyxl.load_workbook("Leveäniemi vattenbalans 301013-3-Update-2026Feb26.xlsx", data_only=False)
wb_v = openpyxl.load_workbook("Leveäniemi vattenbalans 301013-3-Update-2026Feb26.xlsx", data_only=True)
ws_f = wb_f["Process water"]
ws_v = wb_v["Process water"]

# Cu block header is at Excel row 42 (0-based row 41)
# Data starts at row 43 (0-based 42)
print("=== CU BLOCK header rows (Excel 39-42) ===")
for r in range(39, 43):
    for c in range(1, 38):
        v = ws_v.cell(row=r, column=c).value
        f = ws_f.cell(row=r, column=c).value
        if v not in (None,"") or (isinstance(f,str) and f.startswith("=")):
            col_letter = openpyxl.utils.get_column_letter(c)
            print(f"  [{col_letter}{r}] value={v!r}  formula={f!r}")

print("\n=== CU BLOCK first 3 data rows (Excel 43-48) ===")
for r in range(43, 49):
    v_a = ws_v.cell(row=r, column=1).value
    if v_a is None: continue
    print(f"\n  Row {r} (year={v_a})")
    for c in range(1, 38):
        v = ws_v.cell(row=r, column=c).value
        f = ws_f.cell(row=r, column=c).value
        if v not in (None,"") or (isinstance(f,str) and f.startswith("=")):
            col_letter = openpyxl.utils.get_column_letter(c)
            if isinstance(f,str) and f.startswith("="):
                print(f"    [{col_letter}{r}] formula: {f}")
                print(f"           value  : {v}")
            else:
                print(f"    [{col_letter}{r}] value  : {v!r}")

# Also check the Modell During mining sheet for pit pump formula
print("\n\n=== Modell During mining: first 10 rows with values ===")
ws_md_v = wb_v["Modell During mining"]
ws_md_f = wb_f["Modell During mining"]
for r in range(1, 200):
    non_null = []
    for c in range(1, 15):
        v = ws_md_v.cell(row=r, column=c).value
        f = ws_md_f.cell(row=r, column=c).value
        if v not in (None, ""):
            lbl = openpyxl.utils.get_column_letter(c)
            non_null.append(f"[{lbl}{r}]={v!r}")
    if non_null:
        print(f"  Row {r}: {' | '.join(non_null[:10])}")
    if r > 170 and not non_null:
        break
