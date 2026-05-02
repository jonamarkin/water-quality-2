"""
extract_formulas.py
Extract the actual Excel cell formulas from the Process water sheet
to reverse-engineer the consultant's mathematical model.
"""
import openpyxl
import warnings
warnings.filterwarnings("ignore")

EXCEL_FILE = "Leveäniemi vattenbalans 301013-3-Update-2026Feb26.xlsx"

# Load with data_only=False to get formulas, not values
wb_formulas = openpyxl.load_workbook(EXCEL_FILE, data_only=False)
wb_values   = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

ws_f = wb_formulas["Process water"]
ws_v = wb_values["Process water"]

print("=== FORMULA EXTRACTION: Process water sheet ===")
print("(row numbers are 1-based Excel rows)\n")

# The Cu block header is at our row 41 (0-based) = Excel row 42
# We want to look at rows 42-76 (Excel), i.e. rows covering the Cu block header + data rows
# The key columns to inspect are D through AF (cols 4-32 in Excel = cols 3-31 in 0-based)

# Focus: Excel rows 42 to 75 (Cu block), columns A to AM
# Print every non-empty formula cell

BLOCKS_EXCEL = {
    "Cu":  (42, 76),    # Excel row range (1-based, inclusive)
    "NH4": (81, 115),
    "Cl":  (121, 155),
    "Ni":  (160, 194),
}

COL_LETTERS = {
    1:"A",2:"B",3:"C",4:"D",5:"E",6:"F",7:"G",8:"H",9:"I",10:"J",
    11:"K",12:"L",13:"M",14:"N",15:"O",16:"P",17:"Q",18:"R",19:"S",
    20:"T",21:"U",22:"V",23:"W",24:"X",25:"Y",26:"Z",27:"AA",28:"AB",
    29:"AC",30:"AD",31:"AE",32:"AF",33:"AG",34:"AH",35:"AI",36:"AJ",
}

for param, (r_start, r_end) in BLOCKS_EXCEL.items():
    print(f"\n{'='*70}")
    print(f"  BLOCK: {param}  (Excel rows {r_start} to {r_end})")
    print(f"{'='*70}")

    # First print the header row (2 rows before data starts usually)
    for look_row in range(max(1, r_start-3), r_start+1):
        for col in range(1, 38):
            cell_f = ws_f.cell(row=look_row, column=col)
            cell_v = ws_v.cell(row=look_row, column=col)
            if cell_v.value not in (None, ""):
                print(f"  [{COL_LETTERS.get(col,'?')}{look_row}] "
                      f"value={cell_v.value!r}  formula={cell_f.value!r}")

    # Now print first 4 data rows to capture formula patterns
    print(f"\n  --- First 4 data rows (formula patterns) ---")
    data_count = 0
    for excel_row in range(r_start, r_end + 1):
        year_v = ws_v.cell(row=excel_row, column=1).value
        if year_v is None:
            continue
        try:
            yr = float(year_v)
            if not (2013 <= yr <= 2032):
                continue
        except:
            continue

        data_count += 1
        if data_count > 4:
            break

        print(f"\n  Excel row {excel_row}  (year={year_v})")
        for col in range(1, 38):
            cell_f = ws_f.cell(row=excel_row, column=col)
            cell_v = ws_v.cell(row=excel_row, column=col)
            v = cell_v.value
            f = cell_f.value
            if v is None and f is None:
                continue
            lbl = COL_LETTERS.get(col, f"col{col}")
            if isinstance(f, str) and f.startswith("="):
                print(f"    [{lbl}{excel_row}] formula: {f}")
                print(f"           value : {v}")
            elif v is not None:
                print(f"    [{lbl}{excel_row}] value  : {v!r}")
